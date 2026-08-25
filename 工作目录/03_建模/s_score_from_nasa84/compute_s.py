#!/usr/bin/env python3
"""在当前 NASA 84 样本实验上，按历史最终公式计算综合绩效 S。

NASA 来源：regression_task_level/dataset/task_level_table.csv 的 y_nasa
        （与 classification_task_level_nasa 同一套 84 条标签）
步骤来源：data/06_任务表现与操作日志/任务序列完成统计.xlsx
        （黄表头 = 关键子任务；覆盖 26 被试，是现行 84 样本能对齐的唯一全量步骤表）

最终公式（历史 最终S总分计算说明）：
    nasa_reverse = 1 - y_nasa / 10
    weighted_step = 0.75 * key_completion + 0.25 * nonkey_completion
                    （若只有关键或只有非关键，该侧权重归一为 1）
    S = 0.40 * weighted_step + 0.60 * nasa_reverse

S 是绩效分：NASA 越高（负荷越重）→ S 越低。
"""

from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook


HERE = Path(__file__).resolve().parent
REPO = HERE.parents[2]
NASA_TABLE = REPO / "工作目录" / "03_建模" / "regression_task_level" / "dataset" / "task_level_table.csv"
SEQ_FILE = REPO / "data" / "06_任务表现与操作日志" / "任务序列完成统计.xlsx"
OUT_DIR = HERE / "output"

KEY_WEIGHT = 0.75
STEP_WEIGHT_IN_S = 0.40

SHEET_TASK_MAP = {
    "01": "1",
    "02": "2",
    "03": "3",
    "04": "4",
    "05": "5",
    "第二次测试05_06": "5_6",
    "第三次测试05_06": "5_6",
}


def is_yellow_header(cell) -> bool:
    fill = cell.fill
    if fill is None or fill.fill_type != "solid":
        return False
    rgb = str(cell.fill.fgColor.rgb or "").upper()
    return rgb.endswith("FFFF00")


def norm_subject(value: object) -> int | None:
    if value is None or str(value).strip() == "":
        return None
    try:
        return int(float(str(value).strip()))
    except ValueError:
        digits = "".join(ch for ch in str(value) if ch.isdigit())
        return int(digits) if digits else None


def tertile_bin(values: np.ndarray, labels: tuple[str, str, str] = ("低", "中", "高")) -> tuple[np.ndarray, float, float]:
    q_lo, q_hi = np.quantile(values, [1.0 / 3.0, 2.0 / 3.0])
    bins = np.where(values <= q_lo, labels[0], np.where(values <= q_hi, labels[1], labels[2]))
    return bins, float(q_lo), float(q_hi)


def weighted_step(row: pd.Series, key_weight: float = KEY_WEIGHT) -> float:
    n_key = int(row["n_key_steps"])
    n_nonkey = int(row["n_nonkey_steps"])
    if n_key > 0 and n_nonkey > 0:
        return float(key_weight * row["key_completion"] + (1.0 - key_weight) * row["nonkey_completion"])
    if n_key > 0:
        return float(row["key_completion"])
    if n_nonkey > 0:
        return float(row["nonkey_completion"])
    return float("nan")


def read_sequence_steps(path: Path) -> pd.DataFrame:
    wb = load_workbook(path, data_only=True)
    rows: list[dict] = []
    for sheet_name, task in SHEET_TASK_MAP.items():
        if sheet_name not in wb.sheetnames:
            raise FileNotFoundError(f"序列表缺少 sheet：{sheet_name}")
        ws = wb[sheet_name]
        step_cols: list[tuple[int, str, bool]] = []
        for col_idx in range(2, ws.max_column + 1):
            header = ws.cell(1, col_idx).value
            if header is None:
                continue
            step_cols.append((col_idx, str(header), is_yellow_header(ws.cell(1, col_idx))))
        for row_idx in range(2, ws.max_row + 1):
            subject = norm_subject(ws.cell(row_idx, 1).value)
            if subject is None:
                continue
            values, key_values, nonkey_values = [], [], []
            step_detail: dict[str, float] = {}
            for col_idx, header, is_key in step_cols:
                raw = ws.cell(row_idx, col_idx).value
                try:
                    value = float(raw)
                except (TypeError, ValueError):
                    value = np.nan
                if not np.isnan(value):
                    value = 1.0 if value >= 0.5 else 0.0
                step_detail[header] = value
                values.append(value)
                (key_values if is_key else nonkey_values).append(value)
            rows.append(
                {
                    "subject": subject,
                    "task": task,
                    "source_sheet": sheet_name,
                    "n_steps": len(values),
                    "n_key_steps": len(key_values),
                    "n_nonkey_steps": len(nonkey_values),
                    "key_steps": ",".join(h for _, h, k in step_cols if k),
                    "nonkey_steps": ",".join(h for _, h, k in step_cols if not k),
                    "raw_step_completion": float(np.nanmean(values)) if values else np.nan,
                    "key_completion": float(np.nanmean(key_values)) if key_values else np.nan,
                    "nonkey_completion": float(np.nanmean(nonkey_values)) if nonkey_values else np.nan,
                    "step_detail": ";".join(
                        f"{k}={int(v) if v == v else 'NA'}" for k, v in step_detail.items()
                    ),
                }
            )
    steps = pd.DataFrame(rows)
    if steps.duplicated(["subject", "task"]).any():
        dup = steps[steps.duplicated(["subject", "task"], keep=False)]
        raise RuntimeError(f"序列表 (subject, task) 不唯一：\n{dup[['subject','task','source_sheet']]}")
    return steps


def write_report(out: pd.DataFrame, q_nasa: tuple[float, float], q_s: tuple[float, float]) -> str:
    n = len(out)
    n_ok = int(out["has_step_score"].sum())
    s = out["S"].to_numpy(dtype=float)
    nasa = out["y_nasa"].to_numpy(dtype=float)
    rho = float(pd.Series(s).rank().corr(pd.Series(nasa).rank()))
    pearson = float(pd.Series(s).corr(pd.Series(nasa)))

    by_diff = (
        out.groupby("task_difficulty", sort=False)
        .agg(n=("sample_id", "size"), S_mean=("S", "mean"), nasa_mean=("y_nasa", "mean"))
        .reindex(["低", "中", "高"])
    )
    by_nasa_bin = (
        out.groupby("nasa_bin", sort=False)
        .agg(n=("sample_id", "size"), S_mean=("S", "mean"), nasa_mean=("y_nasa", "mean"))
        .reindex(["低", "中", "高"])
    )
    task_order = ["1", "2", "3", "4", "5", "5_6"]
    by_task = (
        out.groupby("task", sort=False)
        .agg(n=("sample_id", "size"), S_mean=("S", "mean"), nasa_mean=("y_nasa", "mean"),
             step_mean=("weighted_step_score", "mean"))
        .reindex(task_order)
        .reset_index()
    )
    cross = pd.crosstab(out["nasa_bin"], out["S_bin"]).reindex(index=["低", "中", "高"], columns=["低", "中", "高"]).fillna(0).astype(int)
    repeats = out[out["sample_id"].astype(str).str.contains("repeat", regex=False)]

    lines: list[str] = []
    lines.append("# 当前 NASA 84 样本上的综合绩效 S\n\n")
    lines.append("在最新 NASA 实验（`classification_task_level_nasa` / `regression_task_level`，84 条被试–任务）上，")
    lines.append("按历史最终公式回算自定义绩效 S。S **不是**新的官方量表，而是 NASA 反向分与关键子任务完成率的加权合成。\n\n")

    lines.append("## 公式\n\n")
    lines.append("```\n")
    lines.append("nasa_reverse      = 1 - y_nasa / 10\n")
    lines.append("weighted_step     = 0.75 * key_completion + 0.25 * nonkey_completion\n")
    lines.append("S                 = 0.40 * weighted_step + 0.60 * nasa_reverse\n")
    lines.append("```\n\n")
    lines.append("- `y_nasa`：当前实验任务级 NASA-TLX 加权总分（范围约 1.33–7.80）。\n")
    lines.append("- 关键子任务：`任务序列完成统计.xlsx` 黄表头列；步骤值 ≥0.5 记 1，否则记 0。\n")
    lines.append("- 若某任务只有关键或只有非关键列，该侧权重归一为 1（任务 5 与 5_6 的列全部为关键）。\n")
    lines.append("- 重复测量（`_repeat_*`）共用该被试–任务的步骤分，但保留各自的 `y_nasa`，因此 S 可以不同。\n\n")

    lines.append("## 覆盖\n\n")
    lines.append(f"- NASA 样本：{n}\n")
    lines.append(f"- 成功配上步骤分：{n_ok} / {n}\n")
    if n_ok < n:
        miss = out.loc[~out["has_step_score"], "sample_id"].tolist()
        lines.append(f"- 未匹配：{miss}\n")
    else:
        lines.append("- 未匹配：无\n")
    lines.append(f"- 重复测量条数：{len(repeats)}\n\n")

    lines.append("## S 分布\n\n")
    lines.append(f"- 范围：[{s.min():.4f}, {s.max():.4f}]\n")
    lines.append(f"- 均值 ± 标准差：{s.mean():.4f} ± {s.std(ddof=0):.4f}\n")
    lines.append(f"- NASA 三分位阈值：低 ≤ {q_nasa[0]:.3f}，中 ≤ {q_nasa[1]:.3f}\n")
    lines.append(f"- S 三分位阈值（绩效高/中/低对应 S 高/中/低）：低 ≤ {q_s[0]:.3f}，中 ≤ {q_s[1]:.3f}\n")
    lines.append(f"- S 与 NASA Spearman ρ = {rho:.3f}（期望为负：负荷越高，绩效 S 越低）\n")
    lines.append(f"- S 与 NASA Pearson r = {pearson:.3f}\n\n")

    lines.append("## 按预设任务难度\n\n")
    lines.append("| 难度 | n | NASA 均值 | S 均值 |\n|---|---:|---:|---:|\n")
    for idx, r in by_diff.iterrows():
        if pd.isna(r["n"]):
            continue
        lines.append(f"| {idx} | {int(r['n'])} | {r['nasa_mean']:.3f} | {r['S_mean']:.3f} |\n")
    low, mid, high = by_diff.loc["低", "S_mean"], by_diff.loc["中", "S_mean"], by_diff.loc["高", "S_mean"]
    mono = (low > mid > high)
    lines.append(f"\nS 是否随难度单调下降（低 > 中 > 高）：{'是' if mono else '否'} ")
    lines.append(f"（{low:.3f} / {mid:.3f} / {high:.3f}）\n\n")

    lines.append("## 按 NASA 三分位档\n\n")
    lines.append("| NASA 档 | n | NASA 均值 | S 均值 |\n|---|---:|---:|---:|\n")
    for idx, r in by_nasa_bin.iterrows():
        lines.append(f"| {idx} | {int(r['n'])} | {r['nasa_mean']:.3f} | {r['S_mean']:.3f} |\n")

    lines.append("\n## NASA 档 × S 档\n\n")
    lines.append("| NASA \\ S | 低 | 中 | 高 |\n|---|---:|---:|---:|\n")
    for idx in ["低", "中", "高"]:
        row = cross.loc[idx]
        lines.append(f"| {idx} | {int(row['低'])} | {int(row['中'])} | {int(row['高'])} |\n")
    agree = float((out["nasa_bin"] == out["S_bin"]).mean())
    # 方向相反：NASA 高应对 S 低。反对角线才是“一致方向”
    flipped = out["nasa_bin"].map({"低": "高", "中": "中", "高": "低"})
    flip_agree = float((flipped == out["S_bin"]).mean())
    lines.append(f"\n- 同名档一致率（低/中/高对低/中/高）：{agree * 100:.1f}% —— 含义有限，因为两指标方向相反\n")
    lines.append(f"- 反向对齐率（NASA 低↔S 高，NASA 高↔S 低，中↔中）：{flip_agree * 100:.1f}%\n\n")

    lines.append("## 按任务类型\n\n")
    lines.append("| task | n | 步骤分均值 | NASA 均值 | S 均值 |\n|---|---:|---:|---:|---:|\n")
    for _, r in by_task.iterrows():
        lines.append(
            f"| {r['task']} | {int(r['n'])} | {r['step_mean']:.3f} | {r['nasa_mean']:.3f} | {r['S_mean']:.3f} |\n"
        )

    lines.append("\n## 重复测量\n\n")
    if len(repeats) == 0:
        lines.append("无。\n")
    else:
        lines.append("这些行与同被试同任务的基准行共用步骤分，S 差异只来自 NASA。\n\n")
        lines.append("| sample_id | subject | task | y_nasa | 步骤分 | S |\n|---|---:|---|---:|---:|---:|\n")
        for _, r in repeats.sort_values(["subject", "task", "sample_id"]).iterrows():
            lines.append(
                f"| {r['sample_id']} | {int(r['subject'])} | {r['task']} | {r['y_nasa']:.3f} | {r['weighted_step_score']:.3f} | {r['S']:.3f} |\n"
            )

    lines.append("\n## 产出文件\n\n")
    lines.append("- `s_score_84samples.csv`：84 行主表（可直接并回 NASA 实验样本）\n")
    lines.append("- `s_components.csv`：步骤组成明细\n")
    lines.append("- `y_s.npy` / `y_nasa.npy` / `sample_s.npy`：与建模目录同序的数组\n")
    lines.append("- `report.md`：本报告\n")
    return "".join(lines)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    print(f"[compute_s] NASA table: {NASA_TABLE}")
    print(f"[compute_s] sequence:   {SEQ_FILE}")

    nasa = pd.read_csv(NASA_TABLE)
    keep = ["sample_id", "subject", "task", "task_difficulty", "n_windows", "y_nasa"]
    missing = [c for c in keep if c not in nasa.columns]
    if missing:
        raise RuntimeError(f"任务级表缺列：{missing}")
    nasa = nasa[keep].copy()
    nasa["subject"] = nasa["subject"].astype(int)
    nasa["task"] = nasa["task"].astype(str)
    nasa["y_nasa"] = nasa["y_nasa"].astype(float)
    nasa["nasa_reverse"] = 1.0 - nasa["y_nasa"] / 10.0

    steps = read_sequence_steps(SEQ_FILE)
    steps["weighted_step_score"] = steps.apply(weighted_step, axis=1)

    merged = nasa.merge(steps, on=["subject", "task"], how="left")
    merged["has_step_score"] = merged["weighted_step_score"].notna()
    merged["S"] = (
        STEP_WEIGHT_IN_S * merged["weighted_step_score"]
        + (1.0 - STEP_WEIGHT_IN_S) * merged["nasa_reverse"]
    )

    nasa_bin, q_lo_n, q_hi_n = tertile_bin(merged["y_nasa"].to_numpy())
    # S 是绩效：高 S = 好。三分位仍按数值低/中/高切，报告里会说明方向。
    s_bin, q_lo_s, q_hi_s = tertile_bin(merged["S"].to_numpy())
    merged["nasa_bin"] = nasa_bin
    merged["S_bin"] = s_bin

    n_miss = int((~merged["has_step_score"]).sum())
    if n_miss:
        print("[compute_s] WARNING unmatched samples:")
        print(merged.loc[~merged["has_step_score"], ["sample_id", "subject", "task"]].to_string(index=False))
    else:
        print(f"[compute_s] all {len(merged)} NASA samples matched a step score")

    main_cols = [
        "sample_id", "subject", "task", "task_difficulty", "n_windows",
        "y_nasa", "nasa_reverse", "nasa_bin",
        "source_sheet", "n_steps", "n_key_steps", "n_nonkey_steps",
        "key_steps", "nonkey_steps",
        "key_completion", "nonkey_completion", "raw_step_completion",
        "weighted_step_score", "S", "S_bin",
    ]
    main_df = merged[main_cols].copy()
    main_df.to_csv(OUT_DIR / "s_score_84samples.csv", index=False, encoding="utf-8-sig")

    comp_cols = main_cols + ["step_detail", "has_step_score"]
    merged[comp_cols].to_csv(OUT_DIR / "s_components.csv", index=False, encoding="utf-8-sig")

    # 与 NASA 建模表同序，方便后续当第二标签用
    np.save(OUT_DIR / "y_s.npy", merged["S"].to_numpy(dtype=np.float64))
    np.save(OUT_DIR / "y_nasa.npy", merged["y_nasa"].to_numpy(dtype=np.float64))
    np.save(OUT_DIR / "sample_s.npy", merged["sample_id"].to_numpy())
    np.save(OUT_DIR / "groups_s.npy", merged["subject"].to_numpy(dtype=np.int64))

    report = write_report(merged, (q_lo_n, q_hi_n), (q_lo_s, q_hi_s))
    (OUT_DIR / "report.md").write_text(report, encoding="utf-8")

    print(f"[compute_s] S range [{merged['S'].min():.4f}, {merged['S'].max():.4f}] "
          f"mean={merged['S'].mean():.4f}")
    print(f"[compute_s] Spearman(S, NASA) = "
          f"{merged['S'].rank().corr(merged['y_nasa'].rank()):.3f}")
    print(f"[compute_s] wrote {OUT_DIR}")


if __name__ == "__main__":
    main()
